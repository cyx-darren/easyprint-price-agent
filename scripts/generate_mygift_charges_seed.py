#!/usr/bin/env python3
"""
Generate SQL seed rows for public.mygift_charges from the MYGIFT XLSX guide.

Usage:
  python3 scripts/generate_mygift_charges_seed.py \
    "/Users/darrenchoong/Downloads/Printing Price Guide-OSSG Gift Rev01.xlsx" \
    supabase/seed/mygift_charges_seed.sql
"""

from __future__ import annotations

import re
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_INPUT = "/Users/darrenchoong/Downloads/Printing Price Guide-OSSG Gift Rev01.xlsx"
DEFAULT_OUTPUT = "supabase/seed/mygift_charges_seed.sql"
SOURCE_DOC = "Printing Price Guide - OSSG Gift Rev01"
SOURCE_REVISION = "Rev01"


def normalize_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value).strip())


def sql_text(value: object | None) -> str:
    if value is None:
        return "null"
    return "'" + str(value).replace("'", "''") + "'"


def sql_int(value: int | None) -> str:
    return "null" if value is None else str(value)


def sql_bool(value: bool) -> str:
    return "true" if value else "false"


def sql_numeric(value: object) -> str:
    amount = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return format(amount, "f")


def sql_text_array(values: list[str]) -> str:
    if not values:
        return "null"
    return "array[" + ", ".join(sql_text(value) for value in values) + "]::text[]"


def parse_quantity(raw_quantity: str) -> tuple[int | None, int | None]:
    text = normalize_text(raw_quantity).lower().replace(",", "")

    def parse_number(token: str) -> int:
        token = token.strip().lower()
        if token.endswith("k"):
            return int(Decimal(token[:-1]) * 1000)
        return int(Decimal(token))

    range_match = re.match(r"^(\d+(?:\.\d+)?k?)\s*-\s*(\d+(?:\.\d+)?k?)$", text)
    if range_match:
        return parse_number(range_match.group(1)), parse_number(range_match.group(2))

    above_match = re.match(r"^(\d+(?:\.\d+)?k?)\s*(?:&|and)?\s*above$", text)
    if above_match:
        return parse_number(above_match.group(1)), None

    return None, None


def classify_print_method(charge_name: str, raw_quantity: str | None = None) -> str:
    name = charge_name.lower()
    quantity = (raw_quantity or "").lower()

    if "uv" in quantity:
        return "uv_print"
    if "engraving" in quantity:
        return "engraving"
    if name.startswith("uv print"):
        return "uv_print"
    if name.startswith("engraving"):
        return "engraving"
    if name.startswith("heat transfer"):
        return "heat_transfer"
    if name.startswith("paper print"):
        return "paper_print"
    if name.startswith("epoxy"):
        return "epoxy"
    return "silkscreen"


def classify_print_spec(print_method: str) -> str | None:
    if print_method == "silkscreen":
        return "1 colour x 1 position"
    if print_method in {"uv_print", "heat_transfer", "paper_print", "epoxy"}:
        return "full colour"
    if print_method == "engraving":
        return "engraving"
    return None


def classify_charge(
    charge_name: str,
    raw_quantity: str,
    charge_amount: object,
    qty_min: int | None,
    print_method: str,
) -> tuple[str, str, str | None]:
    name = charge_name.lower()
    quantity = raw_quantity.lower()
    amount = Decimal(str(charge_amount))

    if "add-on" in quantity or "add-on" in name or "add " in name:
        return "add_on_per_piece", "piece", "Add-on row from MYGIFT workbook."

    if qty_min is None:
        if print_method in {"uv_print", "heat_transfer", "paper_print", "epoxy"}:
            return "per_piece_full_colour", "piece", "Special non-quantity row preserved from MYGIFT workbook."
        return "per_piece", "piece", "Special non-quantity row preserved from MYGIFT workbook."

    if amount >= Decimal("10"):
        return "flat_packet", "packet", None

    if print_method == "silkscreen":
        return "per_piece_per_colour", "piece", None
    if print_method == "engraving":
        return "per_piece", "piece", None
    return "per_piece_full_colour", "piece", None


def expand_code_range(token: str) -> list[str]:
    match = re.fullmatch(r"([A-Z]+)(\d+)-([A-Z]+)?(\d+)", token)
    if not match:
        return [token]

    prefix_start, start_raw, prefix_end, end_raw = match.groups()
    prefix_end = prefix_end or prefix_start
    if prefix_start != prefix_end:
        return [token]

    start = int(start_raw)
    end = int(end_raw)
    if end < start or end - start > 100:
        return [token]

    width = max(len(start_raw), len(end_raw))
    return [f"{prefix_start}{number:0{width}d}" for number in range(start, end + 1)]


def extract_item_codes(charge_name: str) -> list[str]:
    raw_parts = re.findall(r"\(([^)]*)\)", charge_name)
    if charge_name.startswith("*NW17"):
        raw_parts.append("NW17")

    codes: list[str] = []
    seen: set[str] = set()

    for part in raw_parts:
        normalized = re.sub(r"\band\b|\betc\b|Metal", " ", part, flags=re.IGNORECASE)
        tokens = re.findall(r"[A-Z]{1,4}\d*(?:-[A-Z]{0,4}\d+)?|\d{1,4}", normalized.upper())
        last_prefix: str | None = None
        for token in tokens:
            if token.isdigit():
                if last_prefix:
                    token = f"{last_prefix}{token}"
                else:
                    continue

            prefix_match = re.match(r"([A-Z]+)", token)
            if prefix_match:
                last_prefix = prefix_match.group(1)

            for expanded in expand_code_range(token):
                if expanded and expanded not in seen:
                    seen.add(expanded)
                    codes.append(expanded)

    return codes


def make_charge_row(
    *,
    raw_item_name: str,
    raw_quantity: str,
    charge_amount: object,
    notes: str | None = None,
) -> dict[str, object | None]:
    charge_name = normalize_text(raw_item_name)
    raw_quantity_clean = normalize_text(raw_quantity)
    qty_min, qty_max = parse_quantity(raw_quantity_clean)
    if qty_min is None:
        qty_min = 1
        qty_max = None

    print_method = classify_print_method(charge_name, raw_quantity_clean)
    charge_type, unit, inferred_notes = classify_charge(
        charge_name,
        raw_quantity_clean,
        charge_amount,
        parse_quantity(raw_quantity_clean)[0],
        print_method,
    )

    return {
        "vendor": "MYGIFT",
        "charge_name": charge_name,
        "raw_item_name": raw_item_name,
        "item_codes": extract_item_codes(charge_name),
        "print_method": print_method,
        "print_spec": classify_print_spec(print_method),
        "position": None,
        "raw_quantity": raw_quantity_clean,
        "qty_min": qty_min,
        "qty_max": qty_max,
        "charge_type": charge_type,
        "charge_amount": charge_amount,
        "unit": unit,
        "currency": "SGD",
        "gst_included": False,
        "notes": notes or inferred_notes,
        "source_doc": SOURCE_DOC,
        "source_revision": SOURCE_REVISION,
        "effective_date": None,
        "is_active": True,
    }


def parse_workbook(input_path: Path) -> list[dict[str, object | None]]:
    workbook = load_workbook(input_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    rows: list[dict[str, object | None]] = []
    current_item: str | None = None
    item_only_rows: list[str] = []

    for row in worksheet.iter_rows(values_only=True):
        item, quantity, amount = (list(row) + [None, None, None])[:3]

        if item is not None and str(item).strip() and normalize_text(item) != "Item":
            current_item = str(item).strip()

        if quantity is None or not str(quantity).strip() or normalize_text(quantity) == "Quantity":
            if current_item and amount is None and item is not None:
                item_only_rows.append(current_item)
            continue

        if current_item is None:
            raise ValueError(f"Quantity row has no item: {quantity!r}")

        rows.append(
            make_charge_row(
                raw_item_name=current_item,
                raw_quantity=str(quantity).strip(),
                charge_amount=amount,
            )
        )

    for raw_item_name in item_only_rows:
        charge_name = normalize_text(raw_item_name)
        if charge_name.startswith("*NW17 add"):
            rows.append(
                {
                    "vendor": "MYGIFT",
                    "charge_name": charge_name,
                    "raw_item_name": raw_item_name,
                    "item_codes": ["NW17"],
                    "print_method": "silkscreen",
                    "print_spec": "1 colour x 1 position",
                    "position": None,
                    "raw_quantity": "add S$0.05/pc from above price",
                    "qty_min": 1,
                    "qty_max": None,
                    "charge_type": "add_on_per_piece",
                    "charge_amount": Decimal("0.05"),
                    "unit": "piece",
                    "currency": "SGD",
                    "gst_included": False,
                    "notes": "Apply in addition to Non Woven Bag price for NW17.",
                    "source_doc": SOURCE_DOC,
                    "source_revision": SOURCE_REVISION,
                    "effective_date": None,
                    "is_active": True,
                }
            )
        elif charge_name.startswith("*Below is Full Colour Printing Method"):
            rows.append(
                {
                    "vendor": "MYGIFT",
                    "charge_name": charge_name,
                    "raw_item_name": raw_item_name,
                    "item_codes": [],
                    "print_method": None,
                    "print_spec": None,
                    "position": None,
                    "raw_quantity": "note",
                    "qty_min": 0,
                    "qty_max": None,
                    "charge_type": "note",
                    "charge_amount": Decimal("0.00"),
                    "unit": "note",
                    "currency": "SGD",
                    "gst_included": False,
                    "notes": "Full colour printing method section note; metallic gold and silver excluded.",
                    "source_doc": SOURCE_DOC,
                    "source_revision": SOURCE_REVISION,
                    "effective_date": None,
                    "is_active": True,
                }
            )

    return rows


def render_seed_sql(rows: list[dict[str, object | None]], input_path: Path) -> str:
    columns = [
        "vendor",
        "charge_name",
        "raw_item_name",
        "item_codes",
        "print_method",
        "print_spec",
        "position",
        "raw_quantity",
        "qty_min",
        "qty_max",
        "charge_type",
        "charge_amount",
        "unit",
        "currency",
        "gst_included",
        "notes",
        "source_doc",
        "source_revision",
        "effective_date",
        "is_active",
    ]

    values_sql = []
    for row in rows:
        values = [
            sql_text(row["vendor"]),
            sql_text(row["charge_name"]),
            sql_text(row["raw_item_name"]),
            sql_text_array(row["item_codes"]),  # type: ignore[arg-type]
            sql_text(row["print_method"]),
            sql_text(row["print_spec"]),
            sql_text(row["position"]),
            sql_text(row["raw_quantity"]),
            sql_int(row["qty_min"]),  # type: ignore[arg-type]
            sql_int(row["qty_max"]),  # type: ignore[arg-type]
            sql_text(row["charge_type"]),
            sql_numeric(row["charge_amount"]),
            sql_text(row["unit"]),
            sql_text(row["currency"]),
            sql_bool(bool(row["gst_included"])),
            sql_text(row["notes"]),
            sql_text(row["source_doc"]),
            sql_text(row["source_revision"]),
            sql_text(row["effective_date"]),
            sql_bool(bool(row["is_active"])),
        ]
        values_sql.append("  (" + ", ".join(values) + ")")

    return "\n".join(
        [
            "-- Generated by scripts/generate_mygift_charges_seed.py",
            f"-- Source workbook: {input_path}",
            f"-- Row count: {len(rows)}",
            "",
            "insert into public.mygift_charges (",
            "  " + ",\n  ".join(columns),
            ") values",
            ",\n".join(values_sql),
            "on conflict do nothing;",
            "",
        ]
    )


def main() -> None:
    input_path = Path(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_INPUT)
    output_path = Path(sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUTPUT)

    rows = parse_workbook(input_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(render_seed_sql(rows, input_path), encoding="utf-8")

    print(f"Wrote {len(rows)} MYGIFT charge rows to {output_path}")


if __name__ == "__main__":
    main()
